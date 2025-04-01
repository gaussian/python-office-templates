"""
Utility functions for handling Office-related documents (Excel, PowerPoint, etc.).

This module provides helper functions for working with Microsoft Office document formats,
including file type detection and workbook loading utilities.
"""


def get_load_workbook():
    """
    Returns the load_workbook function from openpyxl.

    This function dynamically imports openpyxl's load_workbook function,
    allowing for lazy loading of the dependency only when needed.

    Returns:
        function: The load_workbook function from openpyxl

    Raises:
        ImportError: If openpyxl is not installed in the environment
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError(
            "openpyxl is not installed. Please install it to use this feature."
        )

    return load_workbook
